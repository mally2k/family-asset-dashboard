import openpyxl
import json
import glob
import os
from datetime import datetime

def normalize_maturity(date_val):
    if not date_val or date_val == '-':
        return '-'
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y.%m.%d')
    s = str(date_val).strip()
    parts = s.split('.')
    if len(parts) >= 3:
        y, m, d = parts[0], parts[1], parts[2]
        if len(y) == 2:
            y = '20' + y
        return f"{y}.{int(m):02d}.{int(d):02d}"
    return s

# Find the most recently modified xlsx file
xlsx_files = glob.glob('유주네 자산 현황*.xlsx')
if not xlsx_files:
    print('No xlsx files found')
    exit(0)

# Sort by modification time (most recent first)
xlsx_files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
excel_path = xlsx_files[0]
print(f'Processing: {excel_path}')

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_detail = wb['상세 자산']
ws_loan = wb['대출'] if '대출' in wb.sheetnames else None

# Parse detail sheet
detail_rows = list(ws_detail.iter_rows(values_only=True))
monthly_data = {}
all_details = []

for row in detail_rows[1:]:
    if not row[0]:
        continue
    month = str(row[0]).strip()
    cat = str(row[1] or '').strip()
    owner = str(row[2] or '').strip()
    name = str(row[3] or '').strip()
    amount = float(row[4]) if row[4] is not None else 0
    ret = row[5]
    if ret in ('-', '—', '', None):
        ret = None
    elif isinstance(ret, str):
        try:
            ret = float(ret)
        except (ValueError, TypeError):
            ret = None
    purpose = str(row[6] or '').strip()
    maturity = str(row[7]).strip() if row[7] is not None else '-'
    
    period = month.replace('2026년 ', '2026-')
    period = period.replace('3월', '03').replace('4월', '04').replace('5월', '05').replace('6월', '06')
    period = period.replace('7월', '07').replace('8월', '08').replace('9월', '09').replace('10월', '10').replace('11월', '11').replace('12월', '12')
    
    if period not in monthly_data:
        monthly_data[period] = {}
    if cat not in monthly_data[period]:
        monthly_data[period][cat] = 0
    monthly_data[period][cat] += amount
    
    all_details.append({
        'month': month, 'period': period, 'category': cat, 'owner': owner,
        'name': name, 'amount': amount, 'return': ret, 'purpose': purpose,
        'maturity': maturity, 'type': cat
    })

# Add real estate to all periods
for period in monthly_data:
    monthly_data[period]['부동산'] = 2200000000

# Parse loan sheet
loan_total = 349025014
loan_rows = list(ws_loan.iter_rows(values_only=True)) if ws_loan else []
for row in loan_rows[1:]:
    if row[0] == '총액' and row[2] and isinstance(row[2], (int, float)):
        loan_total = int(row[2])
        break

periods = sorted(monthly_data.keys())
latest_period = periods[-1]
latest = monthly_data[latest_period]

cash_total = (latest.get('예금', 0) + latest.get('적금', 0))
total_val = cash_total + latest.get('주식', 0) + latest.get('개인연금', 0) + latest.get('가상화폐', 0) + latest.get('부동산', 0)

allocations = [
    {'cat': 'real_estate', 'label': '부동산', 'amount': latest.get('부동산', 0)},
    {'cat': 'stocks', 'label': '주식', 'amount': latest.get('주식', 0)},
    {'cat': 'pension', 'label': '개인연금', 'amount': latest.get('개인연금', 0)},
    {'cat': 'cash', 'label': '현금/예금', 'amount': cash_total},
    {'cat': 'crypto', 'label': '가상화폐', 'amount': latest.get('가상화폐', 0)},
]

asset_allocation = []
for a in allocations:
    pct = round((a['amount'] / total_val) * 100, 2) if total_val > 0 else 0
    asset_allocation.append({
        'category': a['cat'], 'category_label': a['label'],
        'amount_krw': round(a['amount']), 'percentage': pct
    })
asset_allocation = [a for a in asset_allocation if a['amount_krw'] > 0]
asset_allocation.sort(key=lambda x: x['amount_krw'], reverse=True)

historical_trends = []
for p in periods:
    d = monthly_data[p]
    t = (d.get('예금', 0) + d.get('적금', 0) + d.get('주식', 0) + d.get('개인연금', 0) + d.get('가상화폐', 0) + d.get('부동산', 0))
    historical_trends.append({
        'period': p,
        'net_assets_krw': round(t - loan_total),
        'categories': {
            'cash': round(d.get('예금', 0) + d.get('적금', 0)),
            'stocks': round(d.get('주식', 0)),
            'pension': round(d.get('개인연금', 0)),
            'crypto': round(d.get('가상화폐', 0)),
            'real_estate': round(d.get('부동산', 0))
        }
    })

# Detail breakdown for latest period
detail_map = {
    'real_estate': [{'name': '아파트', 'amount': 2200000000, 'owner': '공동', 'purpose': '부동산', 'return': None, 'maturity': '-'}],
    'debt': [], 'stocks': [], 'pension': [], 'cash': [], 'crypto': []
}

for d in all_details:
    if d['period'] != latest_period:
        continue
    item = {
        'name': d['name'], 'amount': round(d['amount']), 'owner': d['owner'],
        'purpose': d['purpose'], 'return': d['return'], 'maturity': d['maturity'],
        'type': d['type']
    }
    if d['category'] == '주식':
        detail_map['stocks'].append(item)
    elif d['category'] == '개인연금':
        detail_map['pension'].append(item)
    elif d['category'] in ('예금', '적금'):
        detail_map['cash'].append(item)
    elif d['category'] == '가상화폐':
        detail_map['crypto'].append(item)

# Parse loan details
if ws_loan:
    for row in loan_rows[1:]:
        if not row[0] or row[0] == '총액':
            continue
        if row[2] and isinstance(row[2], (int, float)):
            interest_rate = None
            if row[3] is not None:
                try:
                    interest_rate = float(row[3])
                except (ValueError, TypeError):
                    interest_rate = None
            maturity = normalize_maturity(row[5])
            detail_map['debt'].append({
                'name': str(row[1] or ''),
                'amount': round(row[2]),
                'owner': str(row[0] or ''),
                'purpose': '대출',
                'return': interest_rate,
                'maturity': maturity
            })

result = {
    'total_valuation_krw': round(total_val),
    'net_valuation_krw': round(total_val - loan_total),
    'total_debt_krw': round(loan_total),
    'currency': 'KRW',
    'asset_allocation': asset_allocation,
    'historical_trends': historical_trends,
    'detail_breakdown': detail_map,
    'metadata': {'source': excel_path, 'last_updated': latest_period}
}

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f'data.json updated with {latest_period} data')
